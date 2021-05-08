import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from Program.DataHandler import DataHandler
from Program.GraphEditor import GraphEditor
from Program.KalmanFilter import KalmanFilter
from Program.MySQLClient import MySQLClient

if __name__ == '__main__':
    sql_client = MySQLClient("xxx.xxx.xxx.xxx", "login", "pwd", "dbname")

    value = 'Dew Point °C'
    period = '19 Jan - 19 Feb 2020'

    records = sql_client.get_all_info_by_stations()
    index = DataHandler.get_index_of_value('Dew Point')
    lists_of_measurements = DataHandler.get_exact_value_from_many_my_sql_records(records, index)
    lists_of_measurements = DataHandler.get_lists_of_floats(lists_of_measurements)

    error_in_estimate = 1.0
    error_in_measurement = 1.0
    lists_of_measurements = DataHandler.replace_none_with_dash(lists_of_measurements)

    lists_of_measurements = DataHandler.fill_missing_data_in_lists(lists_of_measurements)

    station_codes = DataHandler.get_station_codes()

    station_codes, lists_of_measurements = DataHandler.zip_codes_and_measurements(station_codes, lists_of_measurements)

    lists_of_estimates = KalmanFilter.get_lists_of_estimates(lists_of_measurements, error_in_estimate,
                                                             error_in_measurement)

    for station_code, measurements, estimates in zip(station_codes, lists_of_measurements, lists_of_estimates):
        graph = GraphEditor(estimates, measurements, value, station_code, period)
        graph.create_est_and_meas_plot()
        graph.save_plot()
        graph.show_plot()

    wb_to_write = Workbook()
    wb_with_accuracies = load_workbook('../../../test_accuracies.xlsx')
    ws = wb_with_accuracies.active

    accuracies = []

    i = 1
    while True:
        if ws['B{}'.format(i)].value is None:
            break
        accuracies.append(float(ws['B{}'.format(i)].value))
        i += 1

    plot_num = 0
    for station_code, measurements, estimates, accuracy in zip(station_codes, lists_of_measurements, lists_of_estimates,
                                                               accuracies):
        ws = wb_to_write.create_sheet(station_code)
        i = 2
        image_name = 'Plot{}.png'.format(plot_num)

        for measurement, estimate in zip(measurements, estimates):
            ws['A{}'.format(i)] = measurement
            ws['B{}'.format(i)] = estimate
            i += 1
        else:
            ws['A1'] = 'Measurements'
            ws['B1'] = 'Estimates'
            ws['AG1'] = 'Accuracy'
            ws['AG2'] = accuracy
            image = openpyxl.drawing.image.Image(image_name)
            image.anchor = 'C1'
            ws.add_image(image)
            plot_num += 1

    else:
        wb_to_write.save('results.xlsx')
